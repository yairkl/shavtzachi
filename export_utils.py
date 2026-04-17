import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from collections import defaultdict
from datetime import datetime

# Palette of soft, readable colors for divisions
DIVISION_COLORS = [
    "E3F2FD", "E8F5E9", "FFF3E0", "F3E5F5", "F1F8E9",
    "FFFDE7", "EFEBE9", "F9FBE7", "E0F2F1", "E1F5FE",
    "FCE4EC", "E8EAF6", "E0F7FA", "F1F8E9", "FFF8E1"
]

def get_division_fill(division_id):
    if division_id is None:
        return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    color = DIVISION_COLORS[division_id % len(DIVISION_COLORS)]
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def export_schedule_to_excel(assignments, start_date, end_date):
    """
    assignments: List of dictionaries with:
        soldier_name, post_name, start (datetime), end (datetime), role_id, division_id, role_name
    """
    MAX_COLS = 15
    
    # 1. Organize data and normalize shift keys
    posts_data = defaultdict(lambda: defaultdict(list))
    for a in assignments:
        start = a['start']
        if isinstance(start, str): start = datetime.fromisoformat(start)
        end = a['end']
        if isinstance(end, str): end = datetime.fromisoformat(end)
        
        # Strip microseconds for grouping
        start = start.replace(microsecond=0)
        end = end.replace(microsecond=0)
        
        posts_data[a['post_name']][(start, end)].append(a)

    # 2. Group posts by their unique shift patterns
    pattern_to_posts = defaultdict(list)
    for post_name, shifts_dict in posts_data.items():
        pattern = tuple(sorted(shifts_dict.keys()))
        num_roles = max([a['role_id'] for a in assignments if a['post_name'] == post_name] + [-1]) + 1
        pattern_to_posts[pattern].append({
            'name': post_name,
            'num_roles': num_roles,
            'pattern': pattern
        })

    # Sort patterns by number of shifts (descending)
    sorted_patterns = sorted(pattern_to_posts.keys(), key=lambda p: len(p), reverse=True)

    # 3. Create "Blocks" (same-pattern posts packed to fit MAX_COLS)
    blocks = []
    for pattern in sorted_patterns:
        posts_in_pattern = pattern_to_posts[pattern]
        current_chunk = []
        chunk_width = 1 # Time col
        for p in posts_in_pattern:
            if chunk_width + p['num_roles'] > MAX_COLS and current_chunk:
                blocks.append({'pattern': pattern, 'posts': current_chunk, 'width': chunk_width})
                current_chunk = []
                chunk_width = 1
            current_chunk.append(p)
            chunk_width += p['num_roles']
        if current_chunk:
            blocks.append({'pattern': pattern, 'posts': current_chunk, 'width': chunk_width})

    # 4. Pack Blocks into horizontal Bands
    bands = []
    current_band = []
    current_band_width = 0
    for b in blocks:
        # If block fits in current band (with 1 col spacer)
        if current_band_width + b['width'] > MAX_COLS and current_band:
            bands.append(current_band)
            current_band = []
            current_band_width = 0
        
        current_band.append(b)
        current_band_width += b['width'] + 1 # +1 for spacer
    if current_band:
        bands.append(current_band)

    # 5. Render
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Schedule"
    
    # Styling
    main_header_font = Font(bold=True, size=11)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    current_row = 1
    
    for band in bands:
        col_cursor = 1
        band_max_height = 0
        
        for block in band:
            pattern = block['pattern']
            posts = block['posts']
            
            # Start of block: Render Time column
            time_header = ws.cell(row=current_row, column=col_cursor)
            time_header.value = "Time"
            time_header.font = header_font
            time_header.fill = header_fill
            time_header.alignment = center_align
            time_header.border = border
            ws.merge_cells(start_row=current_row, start_column=col_cursor, end_row=current_row+1, end_column=col_cursor)
            
            for i, (s, e) in enumerate(pattern):
                time_str = f"{s.strftime('%H:%M')}-{e.strftime('%H:%M')}"
                cell = ws.cell(row=current_row + i + 2, column=col_cursor)
                cell.value = time_str
                cell.alignment = center_align
                cell.border = border
            
            col_cursor += 1
            
            # Render Posts in the block
            for p in posts:
                p_name = p['name']
                num_roles = p['num_roles']
                
                # Headers
                if num_roles == 1:
                    # Single role: Vertically merge post name across both header rows
                    header_cell = ws.cell(row=current_row, column=col_cursor)
                    header_cell.value = p_name
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = center_align
                    header_cell.border = border
                    
                    # Ensure the border and fill apply to the bottom cell too after merge
                    bottom_cell = ws.cell(row=current_row + 1, column=col_cursor)
                    bottom_cell.border = border
                    bottom_cell.fill = header_fill
                    
                    ws.merge_cells(start_row=current_row, start_column=col_cursor, end_row=current_row + 1, end_column=col_cursor)
                else:
                    # Multi role: Main header (Post Name) + Sub-header (Qualifications)
                    main_title = ws.cell(row=current_row, column=col_cursor)
                    main_title.value = p_name
                    main_title.font = main_header_font
                    main_title.alignment = center_align
                    main_title.border = border
                    ws.merge_cells(start_row=current_row, start_column=col_cursor, end_row=current_row, end_column=col_cursor + num_roles - 1)
                    
                    role_names = {}
                    for shift_key in pattern:
                        for a in posts_data[p_name][shift_key]:
                            role_names[a['role_id']] = a.get('role_name', f"Role {a['role_id']+1}")
                    
                    for r in range(num_roles):
                        role_cell = ws.cell(row=current_row + 1, column=col_cursor + r)
                        role_cell.value = role_names.get(r, f"Role {r+1}")
                        role_cell.font = header_font
                        role_cell.fill = header_fill
                        role_cell.alignment = center_align
                        role_cell.border = border
                    
                # Body (Assignments)
                for i, shift_key in enumerate(pattern):
                    shift_assignments = posts_data[p_name][shift_key]
                    role_map = {a['role_id']: a for a in shift_assignments}
                    for r in range(num_roles):
                        cell = ws.cell(row=current_row + i + 2, column=col_cursor + r)
                        a = role_map.get(r)
                        if a:
                            cell.value = a['soldier_name']
                            cell.fill = get_division_fill(a.get('division_id'))
                        cell.border = border
                        cell.alignment = center_align
                
                col_cursor += num_roles
            
            band_max_height = max(band_max_height, len(pattern))
            col_cursor += 1 # Spacer between blocks in band
            
        current_row += band_max_height + 3 # spacing
        
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = 0
        column_index = col[0].column
        column_letter = get_column_letter(column_index)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
